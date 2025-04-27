from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
import openpyxl
from .models import VehicleServiceRecord, MaintenanceType, Vehicle
from companyApp.models import Company
from .models import Vehicle, MaintenanceType
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import Vehicle, VehicleServiceRecord, MaintenanceType, EmergencyMaintenanceRecord
from .forms import VehicleForm, VehicleServiceRecordForm, VehicleMileageForm, VehicleExpiryUpdateForm, EmergencyMaintenanceForm
import io
from django.template.loader import get_template
from django.http import HttpResponse
from .models import Vehicle
from django.db.models import Q
from datetime import datetime
from companyApp.models import Company
from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import VehicleServiceRecord, Vehicle

import qrcode
import io
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import VehicleServiceRecord, Vehicle
from accountApp.models import Employee


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .forms import MaintenanceTypeForm, InspectionItemForm, SubInspectionItemForm, ServiceTypeForm
from .models import MaintenanceType, InspectionItem, SubInspectionItem, ServiceType

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import MaintenanceType, VehicleServiceRecord
import openpyxl
from django.http import HttpResponse
from django.http import JsonResponse

from accountApp.models import Employee

@login_required
def create_vehicle(request):
    if request.method == "POST":
        form = VehicleForm(request.POST)
        if form.is_valid():
            vehicle = form.save(commit=False)
            vehicle.user = request.user

            try:
                employee = Employee.objects.get(user=request.user)
                vehicle.company = employee.company
            except Employee.DoesNotExist:
                # fallback for superuser or error
                return redirect('dashboardApp:admin_dashboard')  # or show an error message

            vehicle.save()
            return redirect('vehicleApp:vehicle_list')
    else:
        form = VehicleForm()

    return render(request, 'vehicleApp/create_vehicle.html', {'form': form})

# WVAm6tS8wV

@login_required
def vehicle_list(request):
    if request.user.is_superuser:
        vehicles = Vehicle.objects.all()
        print("User is Admin")
    else:
        try:
            employee = Employee.objects.get(user=request.user)
            vehicles = Vehicle.objects.filter(company=employee.company)
        except Employee.DoesNotExist:
            return redirect('dashboardApp:admin_dashboard')

    # Apply pagination
    paginator = Paginator(vehicles, 10)  # Show 10 vehicles per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'vehicleApp/vehicle_list.html', {'page_obj': page_obj})





@login_required
def update_vehicle_expiry(request, vehicle_id):
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)

    if request.method == "POST":
        form = VehicleExpiryUpdateForm(request.POST, instance=vehicle)
        if form.is_valid():
            form.save()
            return redirect('vehicleApp:vehicle_list')  # Redirect back to the list

    else:
        form = VehicleExpiryUpdateForm(instance=vehicle)

    return render(request, 'vehicleApp/update_vehicle_expiry.html', {'form': form, 'vehicle': vehicle})



@login_required
def delete_vehicle(request, vehicle_id):
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)

    if request.method == "POST":
        vehicle.delete()
        return redirect('vehicleApp:vehicle_list')  # redirect to your vehicle list

    # Optional: Handle invalid access via GET
    return render(request, 'vehicleApp/invalid_delete_attempt.html')  # or redirect somewhere else




@login_required
def get_vehicles_by_company(request):
    company_id = request.GET.get('company_id')
    vehicles = Vehicle.objects.filter(company_id=company_id).values('id', 'registration_number')
    return JsonResponse(list(vehicles), safe=False)




@login_required
def select_vehicle_mileage(request):
    selected_company = None

    if request.method == "POST":
        company_id = request.POST.get('company')

        if request.user.is_superuser and company_id:
            selected_company = Company.objects.get(id=company_id)

        form = VehicleMileageForm(request.POST, user=request.user, selected_company=selected_company)

        if form.is_valid():
            vehicle = form.cleaned_data['vehicle']
            mileage = form.cleaned_data['mileage_at_service']

            maintenance_category = MaintenanceType.objects.filter(
                kilometer__lte=mileage
            ).order_by('-kilometer').first()


            return redirect('vehicleApp:create_service_record_step2', vehicle.id, mileage, maintenance_category.id if maintenance_category else 0)

    else:
        form = VehicleMileageForm(user=request.user)

    return render(request, 'vehicleApp/select_vehicle_mileage.html', {'form': form})




@login_required
def create_service_record_step2(request, vehicle_id, mileage, category_id):
    vehicle = Vehicle.objects.get(id=vehicle_id)
    maintenance_category = MaintenanceType.objects.get(id=category_id) if category_id != 0 else None

    if request.method == "POST":
        form = VehicleServiceRecordForm(request.POST)
        if form.is_valid():
            service_record = form.save(commit=False)
            service_record.mechanic = request.user  # Automatically assign the logged-in user
            service_record.save()
            return redirect('vehicleApp:service_record_list')  # Redirect to list page

    else:
        form = VehicleServiceRecordForm(initial={
            'vehicle': vehicle,
            'mileage_at_service': mileage,
            'maintenance_type': maintenance_category,
        })


    return render(request, 'vehicleApp/create_service_record.html', {'form': form})


from accountApp.models import Employee

@login_required
def service_record_list(request):
    user = request.user
    selected_vehicle_id = request.GET.get('vehicle')

    if user.is_superuser:
        # Admin: show all
        vehicles = Vehicle.objects.all()
        records = VehicleServiceRecord.objects.all()
    else:
        # Get the company via Employee model
        employee = Employee.objects.filter(user=user).first()
        if employee and employee.company:
            company = employee.company
            vehicles = Vehicle.objects.filter(company=company)
            records = VehicleServiceRecord.objects.filter(vehicle__company=company)
        else:
            # If no company assigned, fallback to empty lists
            vehicles = Vehicle.objects.none()
            records = VehicleServiceRecord.objects.none()

    # Optional: further filter by selected vehicle
    if selected_vehicle_id:
        records = records.filter(vehicle_id=selected_vehicle_id)

    return render(request, 'vehicleApp/service_record_list.html', {
        'vehicles': vehicles,
        'records': records,
        'selected_vehicle_id': selected_vehicle_id
    })


@login_required
def generate_vehicle_qr(request, vehicle_id):
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    service_records = VehicleServiceRecord.objects.filter(vehicle=vehicle)

    # Prepare QR code data (vehicle details + service history)
    qr_data = f"Vehicle: {vehicle.registration_number}\n"
    for record in service_records:
        qr_data += f"{record.mileage_at_service} km - {record.maintenance_category}\n"

    # Generate QR Code
    qr = qrcode.make(qr_data)
    buffer = io.BytesIO()
    qr.save(buffer, format="PNG")

    return HttpResponse(buffer.getvalue(), content_type="image/png")


@login_required
def maintenance_category_list(request):
    categories = MaintenanceType.objects.all()
    return render(request, 'vehicleApp/maintenance_category_list.html', {'categories': categories})




# Create Maintenance Type View
@login_required
def create_maintenance_type(request):
    if request.method == "POST":
        form = MaintenanceTypeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('vehicleApp:maintenance_category_list')  # Redirect to type list
    else:
        form = MaintenanceTypeForm()

    return render(request, 'vehicleApp/create_maintenance_category.html', {'form': form})

# Create Inspection Item View
@login_required
def create_inspection_item(request):
    if request.method == "POST":
        form = InspectionItemForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('dashboardApp:admin_dashboard')  # Redirect to inspection item list
    else:
        form = InspectionItemForm()

    return render(request, 'vehicleApp/create_inspection_item.html', {'form': form})

# Create Sub-Inspection Item View
@login_required
def create_sub_inspection_item(request):
    if request.method == "POST":
        form = SubInspectionItemForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('dashboardApp:admin_dashboard') # Redirect to sub-inspection item list
    else:
        form = SubInspectionItemForm()

    return render(request, 'vehicleApp/create_sub_inspection_item.html', {'form': form})

# Create Service Type View
@login_required
def create_service_type(request):
    if request.method == "POST":
        form = ServiceTypeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('dashboardApp:admin_dashboard')  # Redirect to service type list
    else:
        form = ServiceTypeForm()

    return render(request, 'vehicleApp/create_service_type.html', {'form': form})



@login_required
def maintenance_type_report(request):
    selected_type_id = request.GET.get('maintenance_type')
    maintenance_types = MaintenanceType.objects.all()
    records = []

    if selected_type_id:
        maintenance_type = MaintenanceType.objects.get(id=selected_type_id)

        if request.user.is_superuser or request.user.is_staff:
            # Admin: All records
            records = VehicleServiceRecord.objects.filter(
                maintenance_type=maintenance_type
            ).select_related('vehicle__company', 'vehicle__user', 'maintenance_type')
        else:
            # Regular user: Only their own vehicle records
            records = VehicleServiceRecord.objects.filter(
                maintenance_type=maintenance_type,
                vehicle__user=request.user
            ).select_related('vehicle__company', 'vehicle__user', 'maintenance_type')

        # If exporting to Excel
        if 'export' in request.GET:
            return export_to_excel(records, maintenance_type)

    return render(request, 'vehicleApp/maintenance_type_report.html', {
        'maintenance_types': maintenance_types,
        'records': records,
        'selected_type_id': int(selected_type_id) if selected_type_id else None
    })




@login_required
def company_maintenance_report(request):
    selected_company_id = request.GET.get('company')
    selected_type_id = request.GET.get('maintenance_type')

    companies = Company.objects.all()
    maintenance_types = MaintenanceType.objects.all()
    records = []

    if selected_company_id:
        filters = {'vehicle__company__id': selected_company_id}

        if selected_type_id:
            filters['maintenance_type__id'] = selected_type_id

        if not request.user.is_superuser and not request.user.is_staff:
            filters['vehicle__user'] = request.user

        records = VehicleServiceRecord.objects.filter(**filters).select_related(
            'vehicle__company', 'vehicle__user', 'maintenance_type'
        )

        if 'export' in request.GET:
            return export_company_report(records, selected_company_id, selected_type_id)

    return render(request, 'vehicleApp/company_maintenance_report.html', {
        'companies': companies,
        'maintenance_types': maintenance_types,
        'records': records,
        'selected_company_id': int(selected_company_id) if selected_company_id else None,
        'selected_type_id': int(selected_type_id) if selected_type_id else None
    })


def export_company_report(records, company_id, maintenance_type_id):
    from django.utils.text import slugify

    company = Company.objects.get(id=company_id)
    maintenance_type = MaintenanceType.objects.get(id=maintenance_type_id) if maintenance_type_id else None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Company Maintenance Report"

    ws.append([
        "Vehicle", "Created By", "Company", "Mileage",
        "Main Item", "Sub Item", "Service Type", "Mechanic"
    ])

    for record in records:
        ws.append([
            record.vehicle.registration_number,
            record.vehicle.user.username,
            record.vehicle.company.name,
            f"{record.mileage_at_service} km",
            record.main_item.name,
            record.sub_item.name,
            record.service_type.get_code_display(),
            record.mechanic.username if record.mechanic else 'N/A'
        ])

    filename = f"{slugify(company.name)}"
    if maintenance_type:
        filename += f"_{slugify(maintenance_type.get_name_display())}"
    filename += "_report.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



def export_to_excel(records, maintenance_type):
    from django.utils.text import slugify

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maintenance Report"

    # Header row
    ws.append([
        "Vehicle", "Created By", "Company", "Mileage",
        "Main Item", "Sub Item", "Service Type", "Mechanic"
    ])

    for record in records:
        ws.append([
            record.vehicle.registration_number,
            record.vehicle.user.username,
            record.vehicle.company.name,
            f"{record.mileage_at_service} km",
            record.main_item.name,
            record.sub_item.name,
            record.service_type.get_code_display(),
            record.mechanic.username if record.mechanic else 'N/A'
        ])

    # Use the display name for the filename
    filename = f"{slugify(maintenance_type.get_name_display())}_report.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response






def export_to_excel(records, maintenance_type=None, company=None):
    from django.utils.text import slugify

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maintenance Report"

    # Header
    ws.append([
        "Vehicle", "Created By", "Company", "Mileage",
        "Main Item", "Sub Item", "Service Type", "Mechanic"
    ])

    for record in records:
        ws.append([
            record.vehicle.registration_number,
            record.vehicle.user.username,
            record.vehicle.company.name,
            f"{record.mileage_at_service} km",
            record.main_item.name,
            record.sub_item.name,
            record.service_type.get_code_display(),
            record.mechanic.username if record.mechanic else 'N/A'
        ])

    filename_parts = []
    if maintenance_type:
        filename_parts.append(slugify(maintenance_type.get_name_display()))
    if company:
        filename_parts.append(slugify(company.name))

    filename = "_".join(filename_parts or ["maintenance"]) + "_report.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



@login_required
def combined_maintenance_report(request):
    user = request.user
    selected_type_id = request.GET.get('maintenance_type')
    selected_company_id = request.GET.get('company')

    maintenance_types = MaintenanceType.objects.all()
    companies = Company.objects.all()

    records = VehicleServiceRecord.objects.select_related(
        'vehicle__company', 'vehicle__user', 'maintenance_type',
        'main_item', 'sub_item', 'service_type', 'mechanic'
    )

    if user.is_superuser or user.is_staff:
        # ✅ Admins can filter any company
        if selected_company_id and selected_company_id.isdigit():
            records = records.filter(vehicle__company_id=int(selected_company_id))
    else:
        # ✅ Company Admin must only see their company's records
        try:
            employee = Employee.objects.get(user=user)
            records = records.filter(vehicle__company=employee.company)
            companies = [employee.company]  # Show only their company in dropdown
            selected_company_id = str(employee.company.id)  # Force selected company
        except Employee.DoesNotExist:
            records = records.none()  # No records if user is not linked

    if selected_type_id and selected_type_id.isdigit():
        records = records.filter(maintenance_type_id=int(selected_type_id))

    # Excel Export
    if 'export' in request.GET:
        mt = MaintenanceType.objects.get(id=int(selected_type_id)) if selected_type_id and selected_type_id.isdigit() else None
        company = Company.objects.get(id=int(selected_company_id)) if selected_company_id and selected_company_id.isdigit() else None
        return export_to_excel(records, mt, company)

    return render(request, 'vehicleApp/combined_maintenance_report.html', {
        'maintenance_types': maintenance_types,
        'companies': companies,
        'records': records,
        'selected_type_id': int(selected_type_id) if selected_type_id and selected_type_id.isdigit() else None,
        'selected_company_id': int(selected_company_id) if selected_company_id and selected_company_id.isdigit() else None,
    })


@login_required
def emergency_maintenance_create(request):
    if request.method == 'POST':
        form = EmergencyMaintenanceForm(request.POST)
        if form.is_valid():
            emergency_record = form.save(commit=False)  # Don't save yet
            emergency_record.mechanic = request.user    # Assign logged-in user
            emergency_record.save()                     # Now save
            return redirect('vehicleApp:emergency_maintenance_list')   # Redirect after saving
    else:
        form = EmergencyMaintenanceForm()

    return render(request, 'vehicleApp/emergency_maintenance_form.html', {'form': form})

@login_required
def emergency_maintenance_list(request):
    vehicle_list = EmergencyMaintenanceRecord.objects.filter(mechanic = request.user)
    return render(request, 'vehicleApp/emergency_maintenance_list.html', {'vehicle_list': vehicle_list})
